"""
BC Heat Pump Contractor Directory — Excel Generator
Produces a formatted .xlsx from bc_contractors_master.csv
"""

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# ── Paths ──────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH   = os.path.join(SCRIPT_DIR, "bc_contractors_master.csv")
XLSX_PATH  = os.path.join(SCRIPT_DIR, "BC_HeatPump_Contractors_Master.xlsx")

# ── Colour palette ─────────────────────────────────────────────────────────
BLUE_DARK   = "1E3A5F"   # header bg
BLUE_MID    = "2E6DA4"   # accent
BLUE_LIGHT  = "D6E4F0"   # alt row
WHITE       = "FFFFFF"
GREEN_LIGHT = "D5F5E3"   # Verified_Website
YELLOW_LIGHT= "FEF9E7"   # Unverified
RED_LIGHT   = "FDECEA"   # CAUTION

THIN = Side(style='thin', color="B0C4DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Column widths (characters) ─────────────────────────────────────────────
COL_WIDTHS = {
    "Company_Name":        32,
    "Phone":               17,
    "Website":             28,
    "Email":               30,
    "City":                14,
    "Province":             8,
    "Service_Area_Cities": 35,
    "Brands_Installed":    28,
    "Services":            30,
    "Google_Maps_Link":    12,   # narrow — just clickable link
    "Notes":               48,
    "Verification_Status": 20,
}

def status_fill(status: str) -> PatternFill:
    """Return a fill colour based on verification status."""
    if status == "Verified_Website":
        return PatternFill("solid", fgColor=GREEN_LIGHT)
    elif status == "Confirmed_Call":
        return PatternFill("solid", fgColor="D0F0C0")
    elif "CAUTION" in status or status == "Unverified":
        colour = YELLOW_LIGHT if "CAUTION" not in status else RED_LIGHT
        return PatternFill("solid", fgColor=colour.lstrip("#"))
    return PatternFill("solid", fgColor=WHITE)


def make_sheet_directory(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Directory"

    headers = list(COL_WIDTHS.keys())

    # ── Header row ─────────────────────────────────────────────────────────
    header_font = Font(bold=True, color=WHITE, size=11, name="Calibri")
    header_fill = PatternFill("solid", fgColor=BLUE_DARK)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " "))
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = BORDER

    ws.row_dimensions[1].height = 28

    # ── Data rows ──────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        alt = row_idx % 2 == 0
        status = row.get("Verification_Status", "")

        for col_idx, col_name in enumerate(headers, start=1):
            value = row.get(col_name, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)

            # Base fill: alt rows get light blue, else white
            if col_name == "Verification_Status":
                cell.fill = status_fill(status)
            elif alt:
                cell.fill = PatternFill("solid", fgColor=BLUE_LIGHT)
            else:
                cell.fill = PatternFill("solid", fgColor=WHITE)

            # Font
            cell.font = Font(name="Calibri", size=10)

            # Alignment
            if col_name in ("Phone", "Province", "Verification_Status"):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name in ("Notes", "Service_Area_Cities", "Brands_Installed", "Services"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_name == "Google_Maps_Link":
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")

            # Make Google Maps a hyperlink
            if col_name == "Google_Maps_Link" and value and value.startswith("http"):
                cell.hyperlink = value
                cell.value = "Map"
                cell.font  = Font(name="Calibri", size=10, color=BLUE_MID, underline="single")

            # Make website a hyperlink
            if col_name == "Website" and value and not value.startswith("http"):
                url = "https://" + value
                cell.hyperlink = url
                cell.font = Font(name="Calibri", size=10, color=BLUE_MID, underline="single")

            cell.border = BORDER

        ws.row_dimensions[row_idx].height = 32

    # ── Column widths ──────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_name]

    # ── AutoFilter & Freeze Panes ──────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes   = "A2"   # freeze header row

    # ── Print settings ────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.sheet_view.showGridLines = True


def make_sheet_summary(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Summary")

    from collections import Counter
    cities    = Counter(r["City"] for r in rows)
    statuses  = Counter(r["Verification_Status"] for r in rows)
    services  = Counter()
    for r in rows:
        for svc in r.get("Services", "").split("/"):
            s = svc.strip()
            if s:
                services[s] += 1

    # Title
    ws["A1"] = "BC Heat Pump Contractor Directory — Summary"
    ws["A1"].font  = Font(bold=True, size=14, color=BLUE_DARK, name="Calibri")
    ws["A1"].fill  = PatternFill("solid", fgColor=BLUE_LIGHT)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 24

    # Total
    ws["A2"] = "Total Contractors:"
    ws["B2"] = len(rows)
    ws["A2"].font = ws["B2"].font = Font(bold=True, name="Calibri")

    # -- By City --
    ws["A4"] = "By City"
    ws["A4"].font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws["A4"].fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws["B4"] = "Count"
    ws["B4"].font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws["B4"].fill = PatternFill("solid", fgColor=BLUE_DARK)

    for i, (city, count) in enumerate(sorted(cities.items()), start=5):
        ws.cell(row=i, column=1, value=city).font = Font(name="Calibri", size=10)
        ws.cell(row=i, column=2, value=count).font = Font(name="Calibri", size=10)

    city_end = 4 + len(cities)

    # -- Verification Status --
    status_row = city_end + 2
    ws.cell(row=status_row, column=1, value="Verification Status").font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws.cell(row=status_row, column=1).fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws.cell(row=status_row, column=2, value="Count").font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws.cell(row=status_row, column=2).fill = PatternFill("solid", fgColor=BLUE_DARK)

    for i, (status, count) in enumerate(statuses.items(), start=status_row + 1):
        c1 = ws.cell(row=i, column=1, value=status)
        c2 = ws.cell(row=i, column=2, value=count)
        c1.font = c2.font = Font(name="Calibri", size=10)
        c1.fill = c2.fill = status_fill(status)

    # -- Services Offered --
    svc_row = status_row + len(statuses) + 2
    ws.cell(row=svc_row, column=1, value="Services Offered").font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws.cell(row=svc_row, column=1).fill = PatternFill("solid", fgColor=BLUE_DARK)
    ws.cell(row=svc_row, column=2, value="Contractors").font = Font(bold=True, size=11, color=WHITE, name="Calibri")
    ws.cell(row=svc_row, column=2).fill = PatternFill("solid", fgColor=BLUE_DARK)

    for i, (svc, count) in enumerate(sorted(services.items(), key=lambda x: -x[1]), start=svc_row + 1):
        ws.cell(row=i, column=1, value=svc).font = Font(name="Calibri", size=10)
        ws.cell(row=i, column=2, value=count).font = Font(name="Calibri", size=10)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.freeze_panes = "A2"


def make_sheet_qa_flags(wb: Workbook, rows: list[dict]) -> None:
    """Sheet listing entries that need human verification."""
    ws = wb.create_sheet("QA Flags")

    flagged = [
        r for r in rows
        if r.get("Verification_Status", "") == "Unverified"
        or "CAUTION" in r.get("Notes", "")
        or "unconfirmed" in r.get("Notes", "").lower()
        or "no phone" in r.get("Notes", "").lower()
        or "po box" in r.get("Notes", "").lower()
        or "no address" in r.get("Notes", "").lower()
        or "no street" in r.get("Notes", "").lower()
    ]

    ws["A1"] = f"QA Flags — {len(flagged)} entries need human verification"
    ws["A1"].font = Font(bold=True, size=12, color="B22222", name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor="FDECEA")
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 20

    headers = ["Company_Name", "City", "Phone", "Verification_Status", "Notes"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h.replace("_", " "))
        c.font = Font(bold=True, color=WHITE, name="Calibri")
        c.fill = PatternFill("solid", fgColor=BLUE_DARK)
        c.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(flagged, start=3):
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
            c.font = Font(name="Calibri", size=10)
            c.fill = status_fill(row.get("Verification_Status", ""))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
        ws.row_dimensions[row_idx].height = 36

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 17
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A3"


# ── Main ───────────────────────────────────────────────────────────────────
def main():
    # Load CSV
    with open(CSV_PATH, "r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    # Sort: by Province then City then Company
    rows.sort(key=lambda r: (r.get("Province", ""), r.get("City", ""), r.get("Company_Name", "")))

    print(f"Loaded {len(rows)} contractors from CSV")

    wb = Workbook()

    make_sheet_directory(wb, rows)
    make_sheet_summary(wb, rows)
    make_sheet_qa_flags(wb, rows)

    wb.save(XLSX_PATH)
    print(f"Saved Excel file: {XLSX_PATH}")
    print(f"  Sheets: {', '.join(ws.title for ws in wb.worksheets)}")


if __name__ == "__main__":
    main()
